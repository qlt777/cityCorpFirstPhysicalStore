# Part 2 Programming Coursework

# Use openpyxl to create workbook
from openpyxl import Workbook

wb0 = Workbook("AccountLocation.xlsx")
ws0 = wb0.active

# import openpyxl package
import openpyxl as xl

# file to be copied
wb1 = xl.load_workbook("personal.xlsx")
ws1 = wb1.get_sheet_by_name("Sheet")

import openpyxl
 
# file to be pasted into
wb0 = openpyxl.load_workbook("AccountLocation.xlsx")
wb0_sheet = wb0.get_sheet_by_name("Sheet1")

# copy account numbers as list
def copyRange(1,1,1,1000, ws1):
    rangeSelected = []
    
# loops through selected rows
    for a in range(1,1001,1):
        rowSelected=[]
        # loops through selected columns
        for b in range(1,2,1):
            rowSelected.append(ws1.cell(row = a, column = b).value)
        
    rangeSelected.append(rowSelected)
    
    return rangeSelected

# paste account numbers
# copy data from copyRange into new worksheet
def pasteRange(1,1,1,1000, wb0_sheet, rangeSelected):
    #loop through selected rows
    countRow = 0
    #loop through selected column
    for c in range(1,1001,1):
        countCol = 0
        for d in range(1,2,1):
            
            wb0_sheet.cell(row = c, column = d).value = rangeSelected[0][0]
            countCol += 1
        countRow += 1

# paste account numbers into column 1 of new worksheet        
def createData():
    selectedRange = copyRange(1,1,1,1000,ws1) 
    pastingRange = pasteRange(1,1,1,1000,wb0_sheet,rangeSelected) 

# copy postcodes as list
def copyRange(4,1,4,1000,ws1):
    rangeSelected=[]

# loops through selected rows
    for a in range(1,1001,1):
        rowSelected=[]
        # loops through selected column
        for b in range(4,5,1):
            rowSelected.append(sheet.cell(row = a, column = b).value)
        
    rangeSelected.append(rowSelected)
    
    return rangeSelected

# paste postcodes
# copy data from copyRange into column 2 of new worksheet
def pasteRange(2,1,2,1000, wb0_sheet, rangeSelected):
    #loop through selected rows
    countRow = 0
    for c in range(1,1001,1):
        #loop through selected column
        countCol = 0
        for d in range(2,3,1):
            
            wb0_sheet.cell(row = c, column = d).value = rangeSelected[0][0]
            countCol += 1
        countRow += 1

# paste into column 2 of new worksheet        
def createData():
    selectedRange = copyRange(4,1,4,1000,ws1) 
    pastingRange = pasteRange(2,1,2,1000,wb0_sheet,rangeSelected) 
     
# file to be copied for corresponding values 
wb2 = xl.load_workbook("AccountValue.xlsx")
ws2 = wb2.get_sheet_by_name("Sheet")
    
import openpyxl

# new file to be pasted into
wb0 = openpyxl.load_workbook("AccountLocation.xlsx")
wb0_sheet = wb0.get_sheet_by_name("Sheet1")

# copy corresponding values for men, women and children as list
def copyRange(2,2,4,1001, ws2):
    rangeSelected = []
    
# loops through selected rows
    for a in range(2,1001,1):
        rowSelected= []
        # loops through selected columns
        for b in range(2,5,1):
            rowSelected.append(ws1.cell(row = a, column = b).value)
        
    rangeSelected.append(rowSelected)
    
    return rangeSelected

# paste corresponding values
# copy data from copyRange into columns 5-7 of new worksheet
def pasteRange(5,1,7,1000, wb0_sheet, rangeSelected):
    #loop through selected rows
    countRow = 0
    for c in range(1,1001,1):
        #loop through selected columns
        countCol = 0
        for d in range(5,8,1):
            
            wb0_sheet.cell(row = c, column = d).value = rangeSelected[0][0]
            countCol += 1
        countRow += 1

# pasting into columns 5-7 of new worksheet        
def createData():
    selectedRange = copyRange(2,2,4,1001,ws2) 
    pastingRange = pasteRange(5,1,7,1000,wb0_sheet,rangeSelected) 
    
      
# create a list called tosort as a nested list
tosort = list[i,postcode_i]

# import pandas package 
# use package to create list 
import pandas as pd

data = {'i': [row 'i' in range(1,1001,1)], 'postcode_i': [ws1.cell(row = i, column = 4).value]}

df = pd.read_excel('personal.xlsx', column = 'D')
result = [list(df[x].values) for x in df.columns.values]

# sort list by postcodes using command sort()
tosort.sort(key=lambda x: x[1])


# define variable called nonempty to be true 
my_var = "nonempty"
my_var = True

# import csv package 
import csv

# open postcode file 
with open('postcodes.csv') as file_obj:
    
    # create reader object by passing the file 
    reader_obj = csv.reader(file_obj)
      
    # go through each row in the postcodes.csv file 
    for row in reader_obj:
        print(row)

        for row in range(1,1001,1):
            my_var1 = "test"
            my_var1 = True
            

# if length of tosort is 0
if len(tosort) == 0:
    my_var = False

else: 
    my_var = True 

 
# save resulting workbook   
wb0.save(filename = "AccountLocation.xlsx")


